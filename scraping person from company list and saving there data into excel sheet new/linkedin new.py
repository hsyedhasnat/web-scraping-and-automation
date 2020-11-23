from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

def login(driver):
    driver.get('https://linkedin.com/')
    time.sleep(5)

   
    
    driver.find_element_by_xpath('//*[@id="session_key"]').send_keys('hasnathaider332@gmail.com')
    driver.find_element_by_xpath('//*[@id="session_password"]').send_keys('Shah543@@')
    
    driver.find_element_by_xpath('/html/body/main/section[1]/div[2]/form/button').click()
    time.sleep(4)
    print('login ....')
    return driver
    
    
print('start.................................................')
    
Names=[]
Links=[]
    
    
PATH="geckodriver"
driver = webdriver.Firefox(executable_path=PATH)
    
    
wb = openpyxl.open("Comp_Data.xlsx")
sheet = wb.active
    
wb1 = openpyxl.open("Data.xlsx")
sheet1 = wb1.active
    
print(sheet.max_row)
    
print('Reading File..............')
for row in range(sheet.max_row):
    
    cell1 = sheet.cell(row = row+1, column = 1)
    cell2 = sheet.cell(row = row+1, column = 2)

    Names.append(cell1.value)
    #print(cell1.value)
    Links.append(cell2.value)
    #print(cell2.value)

    
print('Done file is Read.......')
#print(len(Links))
#print(len(Names))
Links=Links[1:]
Names=Names[1:]


driver=login(driver)
press = True
for link,name in zip(Links ,Names):
    Person = []
    Person.append(name)
    Person.append(link)
    driver.get(link)
    time.sleep(5)

    try:
        if(press):
            driver.find_element_by_xpath('/html/body/div[8]/aside/div[1]/header').click()
            press = False
    except:
        try:
            if(press):
                driver.find_element_by_xpath('/html/body/div[7]/aside/div[1]/header').click()
                press = False
        except:
            pass
        
    
    try:
        driver.find_element_by_xpath('/html/body/div[8]/div[3]/div/div[3]/div[1]/section/div/div/div[2]/div[2]/div/div/a').click()
    except:
        driver.find_element_by_xpath('/html/body/div[7]/div[3]/div/div[3]/div[1]/section/div/div/div[2]/div[2]/div/div/a').click()
        
    
    time.sleep(5)
    
    URL=driver.current_url+'&origin=FACETED_SEARCH&title=CEO%20OR%20CTO%20OR%20technical%20OR%20data%20OR%20automation'

    driver.get(URL)
    time.sleep(5)

    try:
        result = driver.find_element_by_xpath('/html/body/div[9]/div[4]/div/div[2]/div/div[2]/div/div/div/div/div[1]/h1').text
        print(result)
        if ('no results found' in result.lower() ):
            continue
    except:
        print('Result Found')
        
    while(True):
        print('Start person..................')
        lis = driver.find_elements_by_xpath('/html/body/div[8]/div[3]/div/div[2]/div/div[2]/div/div/div/div/ul/li')
                                             
        print(len(lis))
        i=1
        for li in lis:
            try:
                a = driver.find_element_by_xpath('/html/body/div[8]/div[3]/div/div[2]/div/div[2]/div/div/div/div/ul/li['+str(i)+']/div/div/div[3]/a').get_attribute('href')
                print(i,a)
                Person.append(a)
            except:
                try:
                    a = driver.find_element_by_xpath('/html/body/div[8]/div[3]/div/div[2]/div/div[2]/div/div/div/div/ul/li['+str(i)+']/div/div/div[2]/a').get_attribute('href')
                    print(i,a)
                    Person.append(a)
                except:
                    print('######')
            driver.execute_script("window.scrollTo(0, "+str(i*70)+")")   
            i+=1


        driver.execute_script("return document.body.scrollHeight")
        try: 
            print('Check 2')
            btu=driver.find_element_by_xpath('/html/body/div[8]/div[3]/div/div[2]/div/div[2]/div/div/div/div/div[2]/div/button[2]')
                                             
            if(btu.is_enabled()):             
                btu.click()
            else:
                print('..No more In both')
                break
            time.sleep(3)
            print('Next Page......')
        except:
            try:
                print('Check 2')
                btu=driver.find_element_by_xpath('/html/body/div[8]/div[3]/div/div[2]/div/div[2]/div/div/div/div/div[1]/div/button[2]') 
                if(btu.is_enabled()):             
                    btu.click()
                else:
                    print('..No more In both')
                    break                  
                time.sleep(3)
                print('Next Page......')
    
            except:
                print('No more In both')
                break
    
    sheet1.append(Person)
    wb1.save("Data.xlsx")

print("Done..............................")
    
    
    

    
